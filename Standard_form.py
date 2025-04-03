from tarfile import data_filter
import streamlit as st
import pandas as pd
from datetime import datetime
from streamlit import columns
from pdfrw import PdfReader
from openpyxl import load_workbook
from functions import total_offerings, read_to_columns, stewards_info,extract_fields_from_excel

#Church Data
church_data=read_to_columns('data\church_id.txt')
st.markdown(church_data.to_html(index=False, header=False), unsafe_allow_html=True)

#Stewards data
stewards_data=stewards_info('data\stewards.txt')
st.markdown(stewards_data.to_html(index=False, header=False),unsafe_allow_html=True)

#open csv file for offering
col1,col2 = st.columns([2,2])
with col1:
    Yearly_offering=total_offerings()
    st.markdown(f'<b><h2> Yearly Offering so far (including Gift Aid) is : £{Yearly_offering:.2f}</b></h2>',
                unsafe_allow_html=True)
# Load the data from the Excel file
file_path = r'data/Church-receipts-and-payments-2024.xlsx'
data = pd.ExcelFile(file_path)

# Create a sidebar for navigation
st.sidebar.title("Navigation")
sheet_names = data.sheet_names
selected_sheet = st.sidebar.selectbox("Select a Sheet", sheet_names)


# Function to display the front page
def display_front_page():
    now = datetime.now()
    current_year = now.strftime("%Y")
    st.title("Church Receipts and Payments")
    st.header("The Methodist Church")
    st.subheader("Standard Form of Accounts")
    church_info = read_to_columns('data/church_id.txt')
    st.write(church_info)

    st.markdown(f"For the year ended 31 August {current_year}", unsafe_allow_html=True)
    noc = church_info['church name'].iloc[0]
    # Other church info...

    # Display Stewards Info
    try:
        stewards_data_save = stewards_data_excel('data/stewards.txt')

        minister = stewards_data_save['minister']
        steward1 = stewards_data_save['steward1']
        steward2 = stewards_data_save['steward2']
        steward3 = stewards_data_save['steward3']
        steward4 = stewards_data_save['steward4']
        treasurer = stewards_data_save['treasurer']
        role_df = pd.DataFrame([minister, steward1, steward2, steward3, steward4, treasurer])
    except Exception as e:
            st.error(f"An error occurred while saving data to Excel: {e}")
    return role_df
    # Button to send data to Excel
if st.button('Submit to Excel'):

    try:
        excel_file = r'data/Church-receipts-and-payments-2025.xlsx'
        wb = load_workbook(excel_file)
        ws = wb['P1 Front page']  # Get the active worksheet

        # Write data to specific cells
        ws['C24'] = minister
        ws['C26'] = steward1
        ws['I26'] = steward2
        ws['C27'] = steward3
        ws['I27'] = steward4
        ws['C36'] = treasurer

        # Save the changes
        wb.save(excel_file)

        # Read back the data to confirm it has been written
        saved_data = {'minister': ws['C24'].value,
                      'steward1': ws['C26'].value,
                      'steward2': ws['I26'].value,
                      'steward3': ws['C27'].value,
                      'steward4': ws['I27'].value,
                      'treasurer': ws['C36'].value,
                      }

        # Display the saved data in Streamlit
        st.success("Data saved successfully to Excel!")
        st.write("Saved Data:")
        st.json(saved_data)  # Display the saved data in a readable format

    except Exception as e:
            st.error(f"An error occurred while saving data to Excel: {e}")


