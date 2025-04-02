import streamlit as st
import pandas as pd
import openpyxl as ox
from datetime import datetime
from functions import total_offerings, bank_interest,lettings_sum


def show_accounting():
    # --add standard header--#
    col1, col2 = st.columns([2, 1], vertical_alignment='top')
    with col1:
        current_date = datetime.now().strftime('%d:%m:%Y')
        st.markdown('<b><h1>Methodist Suite of Finances</b></h1>', unsafe_allow_html=True)
        st.header("Data for Standard Form of Accounts files.")
        st.markdown(f"<b><h5>{current_date}</b></h5>", unsafe_allow_html=True)
    with col2:
        st.image('images/Briggswath_Logo.png')
    st.markdown('<hr></hr>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(['P1 Front page', 'P2 R &P page', 'P3 Summ of Orgs'])
    with tab1:
        st.title("This information will be added to Page 1 of the short form account excel file :")

        church_info_data=pd.read_csv('data/church_info.csv')
        st.header("Church info")
        st.dataframe(church_info_data)


    #loading excel sheet
        workbook=ox.load_workbook('data/Church-receipts-and-payments-2025.xlsx')
        sheet=workbook['P1 Front page']

    #write church info data to worksheet

        if st.button('Submit Page 1 info'):
            if not church_info_data.empty:
                entry = church_info_data.iloc[0]

                if pd.notnull(entry['name']):
                    sheet['C11']=entry['name'].replace('Church', '').strip()
                if pd.notnull(entry['circuit']):
                    sheet['C17']=entry['circuit'].replace('Circuit','').strip()
                if pd.notnull(entry['number']):
                    sheet['K17']=entry['number']
                if pd.notnull(entry['charity number']):
                    sheet['K19']=entry['charity number']
                if pd.notnull(entry['gift aid number']):
                    sheet['K21']=entry['gift aid number']
            #save workbook
        workbook.save('data/Church-receipts-and-payments-2025.xlsx')
        st.success("Church information P1 saved to excel !")
        st.markdown('<hr></hr>', unsafe_allow_html=True)

        #Write Postholder data into P1 Front Page
        st.header('Postholder information')
        postholder_info=pd.read_csv('data/stewards.csv')
        st.dataframe(postholder_info)


        if st.button('Submit postholder information'):
            if not postholder_info.empty:
                post=postholder_info.iloc[0]

                if pd.notnull(post['minister']):
                    sheet['C24']=post['minister']
                if pd.notnull(post['steward1']):
                    sheet['C26']=post['steward1']
                if pd.notnull(post['steward2']):
                    sheet['I26']=post['steward2']
                if pd.notnull(post['steward3']):
                    sheet['C27']=post['steward3']
                if pd.notnull(post['steward4']):
                    sheet['I27'] = post['steward4']
                if pd.notnull(post['steward5']):
                    sheet['C28'] = post['steward5']
                if pd.notnull(post['steward6']):
                    sheet['I28'] = post['steward6']
                if pd.notnull(post['steward7']):
                    sheet['C29'] = post['steward7']
                if pd.notnull(post['treasurer']):
                    sheet['C36'] = post['treasurer']
            workbook.save('data/Church-receipts-and-payments-2025.xlsx')
            st.success("Postholder information P1 saved to excel !")
        st.markdown('<hr></hr>', unsafe_allow_html=True)
    with tab2:

        #Streamlit Datafile creation Page 2 Section A
        st.header("Page 2 of the short form account excel file :")
        st.markdown(f"<hr></hr>", unsafe_allow_html=True)
        st.subheader("Section A")
        offerings=pd.read_csv('data/offering.csv')
        lettings=pd.read_csv('data/total_lettings.csv')
        banking=pd.read_csv('data/banking.csv')
        other_income=pd.read_csv('data/other_income.csv')
        total_offering = offerings['amount'].sum()
        total_df = pd.DataFrame({'Total Offerings and tax recovered': [total_offering]})
        total_lettings=lettings['Total Sum'].sum()
        lettings_df=pd.DataFrame({'Lettings':[total_lettings]})
        banking_interest=banking['recorded annual interest'].sum()
        banking_interest_df=pd.DataFrame({'Bank and CFB interest':[banking_interest]})
        total_other=other_income['amount'].sum()
        other_df=pd.DataFrame({'other receipts':[total_other]})
        final_df = pd.concat([total_df,banking_interest_df,lettings_df,other_df],axis=1)
        st.dataframe(final_df.style.format("{:.2f}"))

        # loading excel sheet
        workbook = ox.load_workbook('data/Church-receipts-and-payments-2025.xlsx')
        sheet = workbook['P2 R & P page']

        if st.button('Submit Section A'):
            if not final_df.empty:
                secA=final_df.iloc[0]
                if pd.notnull(secA['Total Offerings and tax recovered']):
                    sheet['G7']=secA['Total Offerings and tax recovered']
                    sheet['J7'] = secA['Total Offerings and tax recovered']
                if pd.notnull(secA['Bank and CFB interest']):
                    sheet['G8']=secA['Bank and CFB interest']
                    sheet['J8'] = secA['Bank and CFB interest']
                if pd.notnull(secA['Lettings']):
                    sheet['G9']=secA['Lettings']
                    sheet['J9'] = secA['Lettings']
                if pd.notnull(secA['other receipts']):
                    sheet['G10']=secA['other receipts']
                    sheet['J10'] = secA['other receipts']
            workbook.save('data/Church-receipts-and-payments-2025.xlsx')
            st.success("P2 : Section A saved to excel !")
        st.markdown('<hr></hr>', unsafe_allow_html=True)

        #Section B page 2 data
        st.subheader('Section B')
        assessment=pd.read_csv('data/assessment.csv')
        donations=pd.read_csv('data/donations.csv')
        repairs_church=pd.read_csv('data/church_maintenance.csv')
        repairs_cottages=pd.read_csv('data/property_expend.csv')
        utilities=pd.read_csv('data/utility.csv')
        misc=pd.read_csv('data/misc_expend.csv')

        #populate new DataFrame for output
        total_assessment=assessment['amount'].sum()
        assessment_df=pd.DataFrame({'Circuit Assessment or Share':[total_assessment]})
        total_donations_in=donations['amount in'].sum()
        total_donations_out=donations['amount out'].sum()
        total_donations=total_donations_out-total_donations_in
        donation_df=pd.DataFrame({'Donations':[total_donations]})
        total_repairs_church=repairs_church['amount'].sum()

        total_repairs_cottages=repairs_cottages['amount paid'].sum()
        total_repairs=total_repairs_church+total_repairs_cottages
        maintenance_df=pd.DataFrame({'Repairs and Maintenance':[total_repairs]})

        total_utilities=utilities['amount'].sum()
        utilities_df=pd.DataFrame({'Utilities (Insurance, Water charges,heating & lighting)':[total_utilities]})

        total_misc=misc['amount'].sum()
        misc_df=pd.DataFrame({'Other Payments':[total_misc]})

        secB_final_df = pd.concat([assessment_df,donation_df,maintenance_df,utilities_df,misc_df], axis=1)
        st.dataframe(secB_final_df.style.format("{:.2f}"))

        # loading excel sheet
        workbook = ox.load_workbook('data/Church-receipts-and-payments-2025.xlsx')
        sheet = workbook['P2 R & P page']

        if st.button('Submit Section B'):
            if not secB_final_df.empty:
                secB=secB_final_df.iloc[0]
                if pd.notnull(secB['Circuit Assessment or Share']):
                    sheet['G15']=secB['Circuit Assessment or Share']
                    sheet['J15'] =secB['Circuit Assessment or Share']
                if pd.notnull(secB['Donations']):
                    sheet['G16']=secB['Donations']
                    sheet['J16'] =secB['Donations']
                if pd.notnull(secB['Repairs and Maintenance']):
                    sheet['G17']=secB['Repairs and Maintenance']
                    sheet['J17'] =secB['Repairs and Maintenance']
                if pd.notnull(secB['Utilities (Insurance, Water charges,heating & lighting)']):
                    sheet['G18']=secB['Utilities (Insurance, Water charges,heating & lighting)']
                    sheet['J18'] =secB['Utilities (Insurance, Water charges,heating & lighting)']
                if pd.notnull(secB['Other Payments']):
                    sheet['G20']=secB['Other Payments']
                    sheet['J20'] =secB['Other Payments']
                st.dataframe(secB_final_df.style.format("{:.2f}"))
            workbook.save('data/Church-receipts-and-payments-2025.xlsx')
            st.success("P2 : Section B saved to excel !")
        st.markdown('<hr></hr>', unsafe_allow_html=True)

        with tab3:
            st.subheader('Holding Page')



